# -*- coding: cp1251 -*-
import os
import sys
from openpyxl import *
import types
import datetime
import time
from funct import build_list, search_row, convertXLStoXLSX


for file_name in os.listdir("./"):
    if file_name.endswith("xls"):
        path = os.getcwd() + '\\' + file_name
        path = path.replace('\'', '\\') + '\\'
        path = path.rstrip('\\')
        print u'Конвертируем файл', path.decode('cp1251')
        if file_name + "x" in os.listdir("./"):
            continue
        convertXLStoXLSX(path)

valstr = [u'6.68-', u'15.1-']
for i in range(1, 52):
    valstr.append(unicode(u'3.') + unicode(i) + u'-')


newwb = Workbook()
newsheet = newwb.worksheets[0]
newwb_row = 1
newwb_col = 1
# ====add thead=====
thead = [u'Имя файла', u'Объект', u'Акт', u'Период отчета', u'Шифр', u'Наименование работ', u'Ед.изм', u'Кол-во',
         u'Итого по позиции']
for i in range(len(thead)):
    newsheet.cell(newwb_row, i + 1).value = thead[i]
newwb_row += 1

for file_name in os.listdir("./"):
    if all([file_name.endswith("xlsx"), file_name[0] != '~', file_name != 'budjet.xlsx', file_name != 'mat.xlsx',
            file_name != 'meh.xlsx', file_name != 'value.xlsx']):
        print u'Разбираем :', file_name.decode('cp1251')
        wb = load_workbook(file_name, data_only=True)
        sheet = wb.worksheets[0]
        rows = sheet.iter_rows(min_row=169, max_row=169, max_col=20)
        # for row in rows:
        #     s = ''
        #     for cell in row:
        #         if cell.value:
        #             print cell.value, type(cell.value)
        #             s += unicode(cell.value)
        #     if s == u'123456789101112':
        #         print row
        rows = build_list(sheet)
        try:
            coordinates = search_row(rows, [u'Объект'])[0]
            obj = rows[coordinates[0]][coordinates[1] + 1]
        except:
            obj = u'Неизвестно'

        try:
            coordinates = search_row(rows, [u'Номер документа'])[0]
            num_doc = rows[coordinates[0] + 2][coordinates[1]]
        except:
            num_doc = u'Неизвестно'
        try:
            coordinates = search_row(rows, [u'Дата составления'])[0]
            dat = rows[coordinates[0] + 2][coordinates[1] + 1]
        except:
            dat = u'Неизвестно'
        for tmp in search_row(rows, valstr):
            # print tmp, rows[tmp[0]]
            newsheet.cell(newwb_row, 5).value = rows[tmp[0]][tmp[1]]
            newsheet.cell(newwb_row, 6).value = rows[tmp[0]][tmp[1]+1]
            newsheet.cell(newwb_row, 7).value = rows[tmp[0]][tmp[1]+2]
            newsheet.cell(newwb_row, 8).value = rows[tmp[0]][tmp[1]+3]

            itog_coord = search_row(rows[tmp[0]:], [u'Итого по позиции:'])
            if itog_coord:
                newsheet.cell(newwb_row, 9).value = rows[tmp[0]+itog_coord[0][0]][-1]
            else:
                for row in rows[tmp[0]:]:
                    types = [type(cell) is float or type(cell) is long for cell in row]
                    if all(types) and len(types) >= 2:
                        # print row
                        newsheet.cell(newwb_row, 9).value = row[1]
                        break

            newsheet.cell(newwb_row, 1).value = file_name.decode('cp1251')
            newsheet.cell(newwb_row, 2).value = obj
            newsheet.cell(newwb_row, 3).value = num_doc
            newsheet.cell(newwb_row, 4).value = dat
            newwb_row += 1

newwb.save('./value.xlsx')
