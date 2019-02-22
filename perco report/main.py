from PyQt5 import QtWidgets, QtCore
from openpyxl import *
import sys
import ui
import fdb
import datetime


def fdb_connect(path='D:\\bakerdb\\1SCD17K#.FDB', u='sysdba', p='masterkey'):
    """Соединение с БД
    Возвращает курсор к БД"""
    try:
        con = fdb.connect(dsn=path, user=u, password=p, charset='UTF8', fb_library_name='.\\fbclient.dll')
    except:
        input("Couldn't connect to database")
    else:
        cur = con.cursor()

    return cur, con

class Ui(QtWidgets.QMainWindow):

    c, con = fdb_connect()

    def __init__(self):
        super(Ui, self).__init__()
        self.ui = ui.Ui_MainWindow()
        self.ui.setupUi(self)
        self.show()
        self.setWindowTitle(u'PERCO Reporter')

        """ Устанавлеваем период в предидущий месяц """

        now = QtCore.QDate.currentDate()
        ldopm = now.addDays(-now.day())
        self.ui.dateEnd.setDate(ldopm)
        fdopm = ldopm.addDays(-ldopm.day()+1)
        self.ui.dateStart.setDate(fdopm)

        self.ui.btn_report.clicked.connect(self.bul_rep)
        self.ui.btn_rt.clicked.connect(self.change_tabel)
        self.ui.btn_add_staff.clicked.connect(self.add_staff)


        dev = get_divisions(self.c)
        for d in dev:
            self.ui.comboBox.addItem(d[1], int(d[0]))

    def add_staff(self):
        str_lst = self.ui.lst.toPlainText().split('\n')
        # remove empty
        str_lst = list(filter(None, str_lst))
        clean_str_list = list(map(lambda x: x.split('\t'), str_lst))
        # supported format serial.number, LastName
        self.c.execute("select max(s.ID_STAFF) from staff s;")
        res = self.c.fetchall()
        max_staff_id = res[0][0]
        self.c.execute("select max(sc.id_card) from staff_cards sc;")
        res = self.c.fetchall()
        max_card_id = res[0][0]
        sub_div_id = self.ui.comboBox.currentData()

        for new_staff in clean_str_list:
            ind = str(convert_serial_n_to_indifer(new_staff[0]))
            print('tabel', new_staff[0], 'indifer', ind,  'lastname', new_staff[1], 'first', new_staff[2])
            # build sql
            max_staff_id += 1
            max_card_id += 1
            if new_staff[1] == '':
                new_staff[1] = new_staff[0]

            sql_staff_add = f"execute procedure STAFF_$INS({max_staff_id}, '{new_staff[1]}', '{new_staff[2]}','','{new_staff[0]}'," \
                  f"cast('Now' as date),{sub_div_id},0,38740,0,0,cast('Now' as timestamp),0,0,'',null,''," \
                  "'','',null,null,'','','' );"
            # print(sql_staff_add)

            sql_card_add = f"execute procedure STAFF_CARDS_$INS({max_card_id},{max_staff_id},{ind},{ind},0,1,null," \
                           " dateadd(year,2, current_date), 0, 0, 0, 0, 0);"
            # print(sql_card_add)
            self.c.execute(sql_staff_add)
            self.con.commit()
            self.c.execute(sql_card_add)
            self.con.commit()

        self.ui.statusbar.showMessage('Jobs Done.')
        return


    def change_tabel(self):
        staff = get_staff(self.ui.comboBox.currentData(), self.c)
        print(get_cards(self.c, self.con, staff))
        return

    def bul_rep(self):
        self.ui.statusbar.showMessage('Шуршу мозгами........')
        staff = get_staff(self.ui.comboBox.currentData(), self.c)
        start, end = self.ui.dateStart.date(), self.ui.dateEnd.date()

        r = build_report(self.ui.comboBox.currentText(),
                     {'start': QtCore.QDate.toString(start, 'dd.MM.yyyy'),
                      'end': QtCore.QDate.toString(end, 'dd.MM.yyyy')},
                     self.c, staff)
        if r:
            self.ui.statusbar.showMessage('Jobs Done.')



def get_divisions(cursor):
    """ cur - курсор базы данных
    возвращает List [Id, название подразделения]"""

    cursor.execute("select ID_REF, DISPLAY_NAME from SUBDIV_REF ORDER BY DISPLAY_NAME")
    res = cursor.fetchall()

    return res

def get_cards(cursor,con, staff):
    """ cur - курсор базы данных
    возвращает список карт сотрудников"""
    res = []
    for s in staff:
        cursor.execute("select IDENTIFIER from STAFF_CARDS WHERE STAFF_ID = " + str(s[0]))
        r = cursor.fetchall()
        sn = '.'.join(map(str, convert_indifer_to_serial_n(int(r[0][0]))))
        res.append([r[0][0], sn, s[0]])
        # update tabel
        # cursor.execute
        cursor.execute("UPDATE STAFF set TABEL_ID = '" + str(sn) + "' WHERE ID_STAFF = " + str(s[0]))
        con.commit()
    return res

def get_staff(idref, cursor):
    """Получаем список сотрудников для построения отчета"""
    cursor.execute("select "
                   "STAFF_CARDS.STAFF_ID, STAFF_CARDS.IDENTIFIER, STAFF.LAST_NAME, STAFF.TABEL_ID "
                   "FROM STAFF_CARDS, STAFF "
                   "where STAFF_CARDS.STAFF_ID = STAFF.ID_STAFF "
                   "and STAFF_CARDS.STAFF_ID in ("
                   "select STAFF_ID from STAFF_REF WHERE SUBDIV_ID = " + str(idref) + ")")
    staff = cursor.fetchall()
    return staff

def build_report(div, period, cursor, staff = False):
    """ Построение отчета
    Генерирует xls файл
    карта №, карта серия, Фамилия, дата, время, тип прохода"""
    if staff:
        total = 0
        # создаем xl
        new_work_book = Workbook()

        new_sheet = new_work_book.worksheets[0]
        new_work_book_row = 1
        # add table head
        table_head = [u'Identifier', u'Serial.number', u'Last Name', u'Date', u'Time', u'1 In, 2 Out']

        for i in range(len(table_head)):
            new_sheet.cell(new_work_book_row, i + 1).value = table_head[i]
        new_work_book_row += 1

        for s in staff:
            sn, nn = convert_indifer_to_serial_n(s[1])
            cursor.execute("select '" + str(s[1]) + "' , '" + str(sn) + "." + str(nn) +"', '" + str(s[-2]) + "', DATE_PASS,"
                " TIME_PASS, TYPE_PASS, STAFF_ID "
                "from TABEL_INTERMEDIADATE "
                "where DATE_PASS between '" + datetime.datetime.strptime(
                period['start'], '%d.%m.%Y').strftime('%Y-%m-%d') + "' and '" + datetime.datetime.strptime(
                period['end'], '%d.%m.%Y').strftime('%Y-%m-%d') + "' and STAFF_ID = " + str(s[0]))
            res = cursor.fetchall()

            for row in res:
                for i in range(len(row) - 1):
                    new_sheet.cell(new_work_book_row, i + 1).value = row[i]
                new_work_book_row += 1

        fname = str(period['start']) + str(div) + 'report.xlsx'
        new_work_book.save(fname)

    return True


def convert_indifer_to_serial_n(st):
    st = int(st)
    h = hex(st)[2:]
    s = h[0:2]
    n = h[2:]
    return int(s, 16), int(n, 16)


def convert_serial_n_to_indifer(sn):
    snl = sn.split('.')
    s = int(snl[0])
    n = int(snl[1])
    h = ''.join([str(hex(s)), str(hex(n))])
    h = h.replace('0x', '')
    return int(h, 16)


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    window = Ui()
    sys.exit(app.exec_())
