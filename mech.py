# -*- coding: cp1251 -*-
import os
import sys
from openpyxl import *
import types
import datetime
import time
from funct import build_list, search_row, convertXLStoXLSX


for file_name in os.listdir("./"):
    if file_name.endswith("xls"):
        path = os.getcwd() + '\\' + file_name
        path = path.replace('\'', '\\') + '\\'
        path = path.rstrip('\\')
        print u'Конвертируем файл', path.decode('cp1251')
        if file_name + "x" in os.listdir("./"):
            continue
        convertXLStoXLSX(path)

mehstrs = [u'2.1-', u'2.2-', u'2.3-', u'3.29-1905-1', u'3.29-1905-5', u'3.29-1905-8', u'3.29-628-1', u'3.29-628-3',
           u'3.29-1906-1', u'3.29-627-1', u'3.29-627-5', u'3.29-1907-2',u'3.29-627-6',u'3.29-624-3']

newwb = Workbook()
newsheet = newwb.worksheets[0]
newwb_row = 1
newwb_col = 1
# ====add thead=====
thead = [u'Имя файла', u'Объект', u'Акт', u'Период отчета', u'Шифр', u'Наименование работ', u'Ед.изм', u'Кол-во',
         u'ЗП', u'ЭМ', u'в т.ч. ЗПМ']
for i in range(len(thead)):
    newsheet.cell(newwb_row, i + 1).value = thead[i]
newwb_row += 1

for file_name in os.listdir("./"):
    if all([file_name.endswith("xlsx"), file_name[0] != '~', file_name != 'budjet.xlsx', file_name != 'mat.xlsx',
            file_name != 'meh.xlsx', file_name != 'value.xlsx']):
        print u'Разбираем :', file_name.decode('cp1251')
        wb = load_workbook(file_name, data_only=True)
        sheet = wb.worksheets[0]
        rows = build_list(sheet)
        try:
            coordinates = search_row(rows, [u'Объект'])[0]
            obj = rows[coordinates[0]][coordinates[1] + 1]
        except:
            obj = u'Неизвестно'

        try:
            coordinates = search_row(rows, [u'Номер документа'])[0]
            num_doc = rows[coordinates[0] + 2][coordinates[1]]
        except:
            num_doc = u'Неизвестно'
        try:
            coordinates = search_row(rows, [u'Дата составления'])[0]
            dat = rows[coordinates[0] + 2][coordinates[1] + 1]
        except:
            dat = u'Неизвестно'

        for tmp in search_row(rows, mehstrs):
            newsheet.cell(newwb_row, 5).value = rows[tmp[0]][tmp[1]]
            newsheet.cell(newwb_row, 6).value = rows[tmp[0]][tmp[1]+1]
            newsheet.cell(newwb_row, 7).value = rows[tmp[0]][tmp[1]+2]
            newsheet.cell(newwb_row, 8).value = rows[tmp[0]][tmp[1]+3]
            zp = search_row(rows[tmp[0]:tmp[0]+5], [u'ЗП'])
            if zp:
                newsheet.cell(newwb_row, 9).value = rows[tmp[0]+zp[0][0]][-1]
            else:
                newsheet.cell(newwb_row, 9).value = 0
            em = search_row(rows[tmp[0]:tmp[0] + 5], [u'ЭМ'])
            if em:
                newsheet.cell(newwb_row, 10).value = rows[tmp[0] + em[0][0]][-1]
            else:
                newsheet.cell(newwb_row, 10).value = 0
            zpm = search_row(rows[tmp[0]:tmp[0] + 5], [u'в т.ч. ЗПМ'])
            if zpm:
                newsheet.cell(newwb_row, 11).value = rows[tmp[0] + zpm[0][0]][-1]
            else:
                newsheet.cell(newwb_row, 11).value = 0
            newsheet.cell(newwb_row, 1).value = file_name.decode('cp1251')
            newsheet.cell(newwb_row, 2).value = obj
            newsheet.cell(newwb_row, 3).value = num_doc
            newsheet.cell(newwb_row, 4).value = dat
            newwb_row += 1

newwb.save('./meh.xlsx')
