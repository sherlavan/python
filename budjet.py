# -*- coding: cp1251 -*- 
import os
import sys
from openpyxl import *
import codecs
import types
import datetime
import time
from funct import build_list, search_row, convertXLStoXLSX


start_time = time.time()

fcount = 0
# create new xl doc
new_work_book = Workbook()
new_sheet = new_work_book.worksheets[0]
new_work_book_row = 1
new_work_book_col = 1
# add table head
table_head = [u'Имя файла', u'Объект', u'Выполнено', u'Период отчета', u'ЗП К1', u'ЭМ-ЗПМ К1', u'ЗПМ К1',
              u'Материалы К1', u'НР К1', u'СП К1', u'НР и СП К1', u'ЗП К 0,98', u'ЭМ-ЗПМ К 0,98', u'ЗПМ К 0,98',
              u'Материалы К 0,98', u'НР К 0,98', u'СП К 0,98', u'НР и СП К 0,98', u'ЗП Договорная',
              u'ЭМ-ЗПМ Договорная', u'ЗПМ Договорная', u'Материалы Договорная', u'НР Договорная', u'СП Договорная',
              u'НР и СП Договорная']

for i in range(len(table_head)):
    new_sheet.cell(new_work_book_row, i + 1).value = table_head[i]

for file_name in os.listdir("./"):
    if file_name.endswith("xls"):
        path = os.getcwd() + '\\' + file_name
        path = path.replace('\'', '\\') + '\\'
        path = path.rstrip('\\')
        print u'Конвертируем файл', path.decode('cp1251')
        convertXLStoXLSX(path)
new_work_book_row += 1
for file_name in os.listdir("./"):
    if all([file_name.endswith("xlsx"), file_name[0] != '~', file_name != 'budjet.xlsx', file_name != 'mat.xlsx',
            file_name != 'meh.xlsx']):

        wb = load_workbook(file_name)
        sheet = wb.worksheets[0]
        obj = ''
        per = ''
        flag = ''
        work_type = ''
        found_data = False
        jobs_done = 'XZ'
        data37 = []
        fcount += 1
        print u'parsing', file_name.decode('cp1251')
        rows = build_list(sheet)

        new_sheet.cell(new_work_book_row, 1).value = file_name.decode('cp1251')
        coordinates = search_row(rows, [u'Объект'])[0]
        new_sheet.cell(new_work_book_row, 2).value = rows[coordinates[0]][coordinates[1] + 1]

        coordinates = search_row(rows, [u"Выполнено в", u"Работы выполнялись", u'Работы выполнены в'])
        if coordinates:
            new_sheet.cell(new_work_book_row, 3).value = rows[coordinates[0][0]][coordinates[0][1]]
        else:
            new_sheet.cell(new_work_book_row, 3).value = u'Неизвестно'

        try:
            coordinates = search_row(rows, [u'Дата составления'])[0]
            new_sheet.cell(new_work_book_row, 4).value = rows[coordinates[0] + 2][-1]
        except:
            new_sheet.cell(new_work_book_row, 4).value = u'Неизвестно'
        # Зарплата основных рабочих
        coordinates = search_row(rows, [u'Зарплата основных рабочих'])
        try:
            new_sheet.cell(new_work_book_row, 5).value = rows[coordinates[0][0]][coordinates[0][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 5).value = u'Неизвестно'
        try:
            new_sheet.cell(new_work_book_row, 12).value = rows[coordinates[1][0]][coordinates[1][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 12).value = u'Неизвестно'
        # EMZPM1 = EM1 - ZPM1
        # в т.ч. зарплата машинистов
        coordinates = search_row(rows, [u'в т.ч. зарплата машинистов'])
        try:
            new_sheet.cell(new_work_book_row, 7).value = rows[coordinates[0][0]][coordinates[0][1] + 1]
            ZPM = rows[coordinates[0][0]][coordinates[0][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 7).value = u'Неизвестно'
            ZPM = 0
        try:
            new_sheet.cell(new_work_book_row, 14).value = rows[coordinates[1][0]][coordinates[1][1] + 1]
            ZPM09 = rows[coordinates[1][0]][coordinates[1][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 14).value = u'Неизвестно'
            ZPM09 = 0

        coordinates = search_row(rows, [u'Эксплуатация машин'])
        try:
            new_sheet.cell(new_work_book_row, 6).value = rows[coordinates[0][0]][coordinates[0][1] + 1] - ZPM
        except:
            new_sheet.cell(new_work_book_row, 6).value = u'Неизвестно'
        try:
            new_sheet.cell(new_work_book_row, 13).value = rows[coordinates[1][0]][coordinates[1][1] + 1] - ZPM09
        except:
            new_sheet.cell(new_work_book_row, 13).value = u'Неизвестно'
        # //em-zpm
        # materials
        coordinates = search_row(rows, [u'Материалы'])
        try:
            new_sheet.cell(new_work_book_row, 8).value = rows[coordinates[0][0]][coordinates[0][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 8).value = u'Неизвестно'
        try:
            new_sheet.cell(new_work_book_row, 15).value = rows[coordinates[1][0]][coordinates[1][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 15).value = u'Неизвестно'

        # Накладные расходы от ЗП
        coordinates = search_row(rows, [u'Накладные расходы от ЗП'])
        try:
            new_sheet.cell(new_work_book_row, 9).value = rows[coordinates[0][0]][coordinates[0][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 9).value = u'Неизвестно'
        try:
            new_sheet.cell(new_work_book_row, 16).value = rows[coordinates[1][0]][coordinates[1][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 16).value = u'Неизвестно'
        # Сметная прибыль от ЗП
        coordinates = search_row(rows, [u'Сметная прибыль от ЗП'])
        try:
            new_sheet.cell(new_work_book_row, 10).value = rows[coordinates[0][0]][coordinates[0][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 10).value = u'Неизвестно'
        try:
            new_sheet.cell(new_work_book_row, 17).value = rows[coordinates[1][0]][coordinates[1][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 17).value = u'Неизвестно'
        # НР и СП от ЗПМ
        coordinates = search_row(rows, [u'НР и СП от ЗПМ'])
        try:
            new_sheet.cell(new_work_book_row, 11).value = rows[coordinates[-2][0]][coordinates[-2][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 11).value = u'Неизвестно'
        try:
            new_sheet.cell(new_work_book_row, 18).value = rows[coordinates[-1][0]][coordinates[-1][1] + 1]
        except:
            new_sheet.cell(new_work_book_row, 18).value = u'Неизвестно'
        # =L2-E2*0,05

        new_sheet.cell(new_work_book_row, 19).value = '=L' + str(new_work_book_row) + '-E' + str(
            new_work_book_row) + '*0.05'
        new_sheet.cell(new_work_book_row, 20).value = '=M' + str(new_work_book_row) + '-F' + str(
            new_work_book_row) + '*0.05'
        new_sheet.cell(new_work_book_row, 21).value = '=N' + str(new_work_book_row) + '-G' + str(
            new_work_book_row) + '*0.05'
        new_sheet.cell(new_work_book_row, 22).value = '=H' + str(new_work_book_row) + '*0.05'

        new_sheet.cell(new_work_book_row, 23).value = '=P' + str(new_work_book_row) + '-I' + str(
            new_work_book_row) + '*0.05'
        new_sheet.cell(new_work_book_row, 24).value = '=Q' + str(new_work_book_row) + '-J' + str(
            new_work_book_row) + '*0.05'
        new_sheet.cell(new_work_book_row, 25).value = '=R' + str(new_work_book_row) + '-K' + str(
            new_work_book_row) + '*0.05'

        new_work_book_row += 1

new_work_book.save('./budjet.xlsx')
print fcount, 'files parsed for ',
print("--- %s seconds --- " % (time.time() - start_time))
