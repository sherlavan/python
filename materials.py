# -*- coding: cp1251 -*- 
import os
import sys
from openpyxl import *
import types
import datetime
import time
from funct import build_list, search_row, convertXLStoXLSX


for file_name in os.listdir("./"):
    if file_name.endswith("xls"):
        path = os.getcwd() + '\\' + file_name
        path = path.replace('\'', '\\') + '\\'
        path = path.rstrip('\\')
        print u'Конвертируем файл', path.decode('cp1251')
        if file_name + "x" in os.listdir("./"):
            continue
        convertXLStoXLSX(path)
# index for search
index_strings = []
for i in range(27):
    index_strings.append(unicode(u'1.') + unicode(i) + u'-')
index_strings.append(unicode(u'МКЭ-'))
# MAIN
new_work_book = Workbook()
new_sheet = new_work_book.worksheets[0]
new_work_book_row = 1
new_work_book_col = 1
# add table_head
table_head = [u'Имя файла', u'Объект', u'Акт', u'Период отчета', u'Шифр', u'Наименование работ', u'Ед.изм', u'Кол-во',
         u'ВСЕГО в текущих']
for i in range(len(table_head)):
    new_sheet.cell(new_work_book_row, i + 1).value = table_head[i]
new_work_book_row += 1
for file_name in os.listdir("./"):
    if all([file_name.endswith("xlsx"), file_name[0] != '~', file_name != 'budjet.xlsx', file_name != 'mat.xlsx',
            file_name != 'meh.xlsx', file_name != 'value.xlsx']):
        print u'Разбираем :', file_name.decode('cp1251')
        wb = load_workbook(file_name, data_only=True)
        sheet = wb.worksheets[0]
        rows = build_list(sheet)

        try:
            coordinates = search_row(rows, [u'Объект'])[0]
            obj = rows[coordinates[0]][coordinates[1] + 1]
        except:
            obj = u'Неизвестно'

        try:
            coordinates = search_row(rows, [u'Номер документа'])[0]
            num_doc = rows[coordinates[0] + 2][coordinates[1]]
        except:
            num_doc = u'Неизвестно'
        try:
            coordinates = search_row(rows, [u'Дата составления'])[0]
            dat = rows[coordinates[0] + 2][coordinates[1] + 1]
        except:
            dat = u'Неизвестно'

        for coordinates in search_row(rows, index_strings):

            for i in range(4):
                new_sheet.cell(new_work_book_row, 5 + i).value = rows[coordinates[0]][coordinates[1] + i]
            if all([type(rows[coordinates[0]][coordinates[1] + 1]) is unicode, rows[coordinates[0]][coordinates[1] + 1].startswith(u'Возврат')]):
                new_sheet.cell(new_work_book_row, 9).value = rows[coordinates[0] + 1][-1]
            else:
                new_sheet.cell(new_work_book_row, 9).value = rows[coordinates[0]][-1]
            new_sheet.cell(new_work_book_row, 1).value = file_name.decode('cp1251')
            new_sheet.cell(new_work_book_row, 2).value = obj
            new_sheet.cell(new_work_book_row, 3).value = num_doc
            new_sheet.cell(new_work_book_row, 4).value = dat
            new_work_book_row += 1

new_work_book.save('./mat.xlsx')
