# -*- coding: utf8 -*-
import sys
import os
import fdb
import win_unicode_console
import datetime
from openpyxl import *

win_unicode_console.enable()

reload(sys)
sys.setdefaultencoding('utf-8')

menu_actions = {}
last_q_result = []
sub_div = {'id': 0, 'name': ''}
period = {'start': 0, 'end': 0}

STAFF = False

def convert_indifer_to_serial_n(st):
    st = int(st)
    h = hex(st).strip('0x')
    s = h[0:2]
    n = h[2:]
    return int(s, 16), int(n, 16)

# Main menu
def main_menu():
    os.system('cls')
    print '\nOrganization is set to : ID = ', sub_div['id'], 'name = ', sub_div['name'], '\n'
    print unicode('Период установлен с: '), period['start'], unicode(' по: '), period['end'], '\n'

    print "Please choose the menu you want to start:"
    print "1. Show divisions"
    print unicode("2. Показать сотрудников и пропуска")
    print "st set date from to"
    print "br build report"
    print "\n0. Quit"
    choice = raw_input(" >>  ")
    exec_menu(choice)

    return

# Execute menu
def exec_menu(choice):
    os.system('cls')
    ch = choice.lower()
    if ch == '':
        menu_actions['main_menu']()
    else:
        try:
            menu_actions[ch]()
        except KeyError:
            print "Invalid selection, please try again.\n"
            menu_actions['main_menu']()
    return


# Menu 1
def show_divisions():
    print "Hello Menu 1 !\n"
    from datetime import timedelta, datetime


    # dt = (datetime.now() - timedelta(minutes=30)).replace(microsecond=0)
    cur.execute("select ID_REF, DISPLAY_NAME from SUBDIV_REF ORDER BY DISPLAY_NAME")

    res = cur.fetchall()
    global last_q_result
    last_q_result = res
    counter = 0
    for row in res:
        counter += 1
        for cell in row:
            print cell,
        print ''
        if counter % 40 == 0:
            raw_input()

    print "oi org. index input"
    print "9. Back"
    print "0. Quit"
    choice = raw_input(" >>  ")
    exec_menu(choice)
    return

def organization_index():
    print unicode('\nВведите ID Организации!')
    ss = int(raw_input('ID:'))
    first_row_in = False
    print ss
    for r in range(len(last_q_result)):
        for c in range(len(last_q_result[r])):
            data = last_q_result[r][c]
            if data == ss:
                first_row_in = last_q_result[r]
                break
        print ''
    if first_row_in:
        global sub_div
        sub_div['id'] = int(first_row_in[0])
        sub_div['name'] = unicode(first_row_in[1])
        print 'organization is set to :', sub_div
    back()


# Menu 2
def show_staff():
    cur.execute("select STAFF_CARDS.STAFF_ID, STAFF_CARDS.IDENTIFIER, STAFF.LAST_NAME, STAFF.TABEL_ID FROM STAFF_CARDS, STAFF where STAFF_CARDS.STAFF_ID = STAFF.ID_STAFF and STAFF_CARDS.STAFF_ID in (select STAFF_ID from STAFF_REF WHERE SUBDIV_ID = " + str(sub_div['id']) + ")")
    res = cur.fetchall()
    global STAFF
    STAFF = res
    for row in res:
        for cell in row:
            print cell,
        print ''

    print "\nololol\n"
    print "9. Back"
    print "0. Quit"
    raw_input(" >> ")
    back()
    return

def set_date():
    print unicode('\nВведите период дат, например 26.01.2019')
    global period
    period['start'] = datetime.datetime.strptime(raw_input('Enter Start date in the format dd.mm.yyyy: '), '%d.%m.%Y')
    period['end'] = datetime.datetime.strptime(raw_input('Enter Start date in the format dd.mm.yyyy: '), '%d.%m.%Y')
    back()
    return

def build_report():
    if STAFF:
        total = 0
        # создаем xl
        new_work_book = Workbook()
        new_sheet = new_work_book.worksheets[0]
        new_work_book_row = 1
        # new_work_book_col = 1
        # add table head
        table_head = [u'Identifier', u'Serial.number', u'Date', u'Time', u'1 In, 2 Out']

        for i in range(len(table_head)):
            new_sheet.cell(new_work_book_row, i + 1).value = table_head[i]
        new_work_book_row += 1

        for s in STAFF:
            sn, nn = convert_indifer_to_serial_n(s[1])
            cur.execute("select '" + str(s[1]) + "' , '" + str(sn) + "." + str(nn) + "', DATE_PASS, TIME_PASS, TYPE_PASS, STAFF_ID "
                "from TABEL_INTERMEDIADATE "
                "where DATE_PASS between '" + datetime.datetime.strptime(
                period['start'], '%d.%m.%Y').strftime('%Y-%m-%d') + "' and '" + datetime.datetime.strptime(
                period['end'], '%d.%m.%Y').strftime('%Y-%m-%d') + "' and STAFF_ID = " + str(s[0]))
            res = cur.fetchall()
            total += len(res)


            for row in res:
                for i in range(len(row) - 1):
                    new_sheet.cell(new_work_book_row, i + 1).value = row[i]
                new_work_book_row += 1

        fname = unicode('./'+str(sub_div['name']) + str(period['start'])+'.xlsx')
        new_work_book.save(fname)

        print 'total records: ', total


    raw_input('waiting enter...')
    back()
    return

# Back to main menu
def back():
    menu_actions['main_menu']()


# Exit program
def exit():
    sys.exit()

# Menu definition
menu_actions = {
    'main_menu': main_menu,
    '1': show_divisions,
    '2': show_staff,
    '9': back,
    '0': exit,
    'oi': organization_index,
    'st': set_date,
    'br': build_report,
    'q': exit,
}


# Main Program
if __name__ == "__main__":
    # Launch main menu
    con = fdb.connect(dsn='D:\\bakerdb\\1SCD17K#.FDB', user='sysdba', password='masterkey', charset='UTF8')
    cur = con.cursor()

    today = datetime.date.today()
    first_day = today.replace(day=1)
    ld_lastMonth = first_day - datetime.timedelta(days=1)
    fd_lastMonth = ld_lastMonth.replace(day=1)
    period['start'] = fd_lastMonth.strftime("%d.%m.%Y")
    period['end'] = ld_lastMonth.strftime("%d.%m.%Y")
    main_menu()
